"""
PROPHET STRATEGIES
Export — generates CSV trade logs and formatted XLSX analysis workbook
"""
from __future__ import annotations
import logging
from pathlib import Path
from typing import Optional

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

from analysis.optimizer import rank_results, best_per_strategy

logger = logging.getLogger(__name__)

# ── Colour palette ────────────────────────────────────────────────────────────
C_DARK_BG     = "0D1117"   # header background
C_ACCENT      = "7C3AED"   # purple accent
C_ACCENT_LITE = "EDE9FE"   # light purple
C_GREEN_BG    = "D1FAE5"
C_RED_BG      = "FEE2E2"
C_GREEN_FG    = "065F46"
C_RED_FG      = "991B1B"
C_YELLOW_BG   = "FEF3C7"
C_HEADER_FONT = "FFFFFF"
C_SUBHEAD     = "374151"
C_BORDER      = "E5E7EB"


def _side(color: str = C_BORDER) -> Side:
    return Side(style="thin", color=color)


def _border() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=C_DARK_BG)


def _accent_fill() -> PatternFill:
    return PatternFill("solid", fgColor=C_ACCENT)


def _lite_fill() -> PatternFill:
    return PatternFill("solid", fgColor=C_ACCENT_LITE)


def _header_font(bold: bool = True) -> Font:
    return Font(name="Calibri", bold=bold, color=C_HEADER_FONT, size=11)


def _title_font() -> Font:
    return Font(name="Calibri", bold=True, color=C_DARK_BG, size=14)


def _subhead_font() -> Font:
    return Font(name="Calibri", bold=True, color=C_SUBHEAD, size=11)


def _body_font() -> Font:
    return Font(name="Calibri", size=10, color="111827")


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _autowidth(ws, min_w: int = 8, max_w: int = 40):
    for col in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0) for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(length + 2, min_w), max_w
        )


def _fmt_pnl(val: float) -> str:
    return f"${val:,.2f}" if val >= 0 else f"(${ abs(val):,.2f})"


class ExportManager:
    """Generates all output files from grid search results."""

    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        (self.output_dir / "csv").mkdir(parents=True, exist_ok=True)
        (self.output_dir / "xlsx").mkdir(parents=True, exist_ok=True)
        (self.output_dir / "reports").mkdir(parents=True, exist_ok=True)

    # ── Public API ────────────────────────────────────────────────────────────

    def export_all(self, results_csv: str) -> dict[str, Path]:
        """
        Given path to master_results.csv, generate:
          - top_results.csv
          - prophet_analysis.xlsx
        Returns dict of output paths.
        """
        df = pd.read_csv(results_csv)
        _coerce_numerics(df)
        logger.info(f"Loaded {len(df):,} results from {results_csv}")

        paths = {}
        paths["top_csv"]  = self._export_top_csv(df)
        paths["xlsx"]     = self._export_xlsx(df)
        return paths

    # ── CSV export ────────────────────────────────────────────────────────────

    def _export_top_csv(self, df: pd.DataFrame) -> Path:
        top = rank_results(df, top_n=200)
        path = self.output_dir / "csv" / "top_results.csv"
        top.to_csv(path, index=False)
        logger.info(f"Top {len(top)} results → {path}")
        return path

    # ── XLSX export ───────────────────────────────────────────────────────────

    def _export_xlsx(self, df: pd.DataFrame) -> Path:
        path = self.output_dir / "xlsx" / "prophet_analysis.xlsx"
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        self._sheet_summary(wb, df)
        self._sheet_top_results(wb, df)
        self._sheet_by_strategy(wb, df)
        self._sheet_by_crypto(wb, df)
        self._sheet_exit_analysis(wb, df)
        self._sheet_fill_model_comparison(wb, df)
        self._sheet_monthly_pnl(wb, df)
        self._sheet_raw(wb, df)

        wb.save(str(path))
        logger.info(f"XLSX → {path}")
        return path

    # ── Sheet: Summary Dashboard ──────────────────────────────────────────────

    def _sheet_summary(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("📊 Summary")
        ws.sheet_view.showGridLines = False

        # Title block
        ws.merge_cells("B2:H2")
        ws["B2"] = "PROPHET STRATEGIES — Performance Summary"
        ws["B2"].font = Font(name="Calibri", bold=True, size=18, color=C_DARK_BG)
        ws["B2"].alignment = _center()

        ws.merge_cells("B3:H3")
        ws["B3"] = "Polymarket Crypto Backtesting System"
        ws["B3"].font = Font(name="Calibri", size=12, color="6B7280", italic=True)
        ws["B3"].alignment = _center()

        # KPI cards row
        row = 5
        kpis = _compute_kpis(df)
        for col_idx, (label, value, fmt) in enumerate(kpis, start=2):
            col = get_column_letter(col_idx)
            # Card background
            fill = _lite_fill() if col_idx % 2 == 0 else PatternFill("solid", fgColor="F3F4F6")
            for r in range(row, row + 3):
                for c in range(col_idx, col_idx + 1):
                    ws.cell(r, c).fill = fill

            ws[f"{col}{row}"] = label
            ws[f"{col}{row}"].font = Font(name="Calibri", size=9, color="6B7280", bold=True)
            ws[f"{col}{row}"].alignment = _center()

            ws[f"{col}{row+1}"] = value
            ws[f"{col}{row+1}"].font = Font(name="Calibri", size=16, bold=True, color=C_DARK_BG)
            ws[f"{col}{row+1}"].alignment = _center()

            ws[f"{col}{row+2}"] = fmt
            ws[f"{col}{row+2}"].font = Font(name="Calibri", size=8, color="9CA3AF")
            ws[f"{col}{row+2}"].alignment = _center()

        # Best configs table
        row = 10
        ws[f"B{row}"] = "Best Configuration per Strategy × Crypto × Fill Model"
        ws[f"B{row}"].font = _title_font()

        row += 1
        best = best_per_strategy(df)
        if not best.empty:
            cols = ["strategy", "crypto", "fill_model", "total_net_pnl", "roi_pct",
                    "sharpe_ratio", "win_rate", "profit_factor", "max_drawdown",
                    "fill_rate", "composite_score"]
            cols = [c for c in cols if c in best.columns]
            headers = [c.replace("_", " ").title() for c in cols]

            for ci, h in enumerate(headers, start=2):
                cell = ws.cell(row, ci, h)
                cell.fill = _header_fill()
                cell.font = _header_font()
                cell.alignment = _center()
                cell.border = _border()
            ws.row_dimensions[row].height = 22

            for ri, (_, r) in enumerate(best.iterrows(), start=1):
                row += 1
                bg = PatternFill("solid", fgColor="F9FAFB") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                for ci, col in enumerate(cols, start=2):
                    val = r.get(col, "")
                    cell = ws.cell(row, ci)
                    cell.border = _border()
                    cell.fill = bg
                    cell.alignment = _center()
                    cell.font = _body_font()

                    if col == "total_net_pnl":
                        cell.value = val if pd.notna(val) else 0
                        cell.number_format = '$#,##0.00;($#,##0.00);"-"'
                        if pd.notna(val):
                            cell.font = Font(name="Calibri", size=10, bold=True,
                                             color=C_GREEN_FG if val >= 0 else C_RED_FG)
                    elif col in ("roi_pct", "win_rate", "fill_rate"):
                        cell.value = (val / 100) if pd.notna(val) else 0
                        cell.number_format = "0.0%"
                    elif col in ("sharpe_ratio", "profit_factor", "composite_score"):
                        cell.value = round(float(val), 4) if pd.notna(val) else 0
                        cell.number_format = "0.0000"
                    elif col == "max_drawdown":
                        cell.value = val if pd.notna(val) else 0
                        cell.number_format = '$#,##0.00'
                    else:
                        cell.value = str(val) if pd.notna(val) else ""

        _autowidth(ws)
        ws.column_dimensions["A"].width = 3  # gutter

    # ── Sheet: Top 50 Results ─────────────────────────────────────────────────

    def _sheet_top_results(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("🏆 Top Results")
        ws.sheet_view.showGridLines = False

        top = rank_results(df, top_n=50)
        if top.empty:
            ws["A1"] = "No results passed minimum filters."
            return

        ws["A1"] = "TOP 50 CONFIGURATIONS — Ranked by Composite Score"
        ws["A1"].font = _title_font()

        # Identify param columns
        param_cols = [c for c in top.columns if c.startswith("param_")]
        metric_cols = ["total_net_pnl", "roi_pct", "sharpe_ratio", "win_rate",
                       "profit_factor", "max_drawdown", "fill_rate", "filled_trades",
                       "composite_score"]
        id_cols = ["strategy", "crypto", "fill_model"]
        all_cols = id_cols + metric_cols + param_cols
        all_cols = [c for c in all_cols if c in top.columns]

        row = 3
        headers = [c.replace("param_", "").replace("_", " ").title() for c in all_cols]
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row, ci, h)
            cell.fill = _accent_fill()
            cell.font = _header_font()
            cell.alignment = _center()
            cell.border = _border()
        ws.row_dimensions[row].height = 22
        ws.freeze_panes = f"A{row+1}"

        for ri, (_, r) in enumerate(top.iterrows(), start=1):
            row += 1
            bg = PatternFill("solid", fgColor="F5F3FF") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for ci, col in enumerate(all_cols, start=1):
                cell = ws.cell(row, ci)
                cell.border = _border()
                cell.fill = bg
                cell.alignment = _center()
                cell.font = _body_font()
                val = r.get(col)

                if col == "total_net_pnl":
                    cell.value = float(val) if pd.notna(val) else 0
                    cell.number_format = '$#,##0.00;($#,##0.00);"-"'
                    if pd.notna(val):
                        cell.fill = PatternFill("solid", fgColor=C_GREEN_BG if val >= 0 else C_RED_BG)
                        cell.font = Font(name="Calibri", size=10, bold=True,
                                         color=C_GREEN_FG if val >= 0 else C_RED_FG)
                elif col in ("roi_pct", "win_rate", "fill_rate"):
                    cell.value = (float(val) / 100) if pd.notna(val) else 0
                    cell.number_format = "0.0%"
                elif col in ("sharpe_ratio", "profit_factor", "composite_score"):
                    cell.value = round(float(val), 4) if pd.notna(val) else 0
                    cell.number_format = "0.0000"
                elif col == "max_drawdown":
                    cell.value = float(val) if pd.notna(val) else 0
                    cell.number_format = '$#,##0.00'
                elif col == "filled_trades":
                    cell.value = int(val) if pd.notna(val) else 0
                    cell.number_format = "#,##0"
                else:
                    cell.value = str(val) if pd.notna(val) else ""

        # Composite score color scale
        score_col = all_cols.index("composite_score") + 1 if "composite_score" in all_cols else None
        if score_col:
            col_letter = get_column_letter(score_col)
            ws.conditional_formatting.add(
                f"{col_letter}4:{col_letter}{row}",
                ColorScaleRule(
                    start_type="min", start_color="FEE2E2",
                    mid_type="percentile", mid_value=50, mid_color="FEF3C7",
                    end_type="max", end_color="D1FAE5",
                )
            )

        _autowidth(ws)

    # ── Sheet: By Strategy ────────────────────────────────────────────────────

    def _sheet_by_strategy(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("📈 By Strategy")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "PERFORMANCE BY STRATEGY"
        ws["A1"].font = _title_font()

        grouped = _group_summary(df, "strategy")
        _write_summary_table(ws, grouped, start_row=3)
        _autowidth(ws)

    # ── Sheet: By Crypto ──────────────────────────────────────────────────────

    def _sheet_by_crypto(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("₿ By Crypto")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "PERFORMANCE BY CRYPTO ASSET"
        ws["A1"].font = _title_font()

        grouped = _group_summary(df, "crypto")
        _write_summary_table(ws, grouped, start_row=3)
        _autowidth(ws)

    # ── Sheet: Exit Strategy Analysis ─────────────────────────────────────────

    def _sheet_exit_analysis(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("🎯 Exit Analysis")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "EXIT STRATEGY ANALYSIS"
        ws["A1"].font = _title_font()

        if "param_exit_strategy" not in df.columns:
            ws["A3"] = "No exit_strategy param found in results."
            return

        grouped = _group_summary(df, "param_exit_strategy")
        # Sort by the multiplier order
        exit_order = [
            "hold_to_resolution",
            "sell_at_2x", "sell_at_5x", "sell_at_10x", "sell_at_15x",
            "sell_at_25x", "sell_at_50x", "sell_at_75x", "sell_at_100x",
            "sell_at_125x", "sell_at_150x",
        ]
        grouped["_sort"] = grouped["param_exit_strategy"].apply(
            lambda x: exit_order.index(x) if x in exit_order else 99
        )
        grouped = grouped.sort_values("_sort").drop(columns="_sort")

        _write_summary_table(ws, grouped, start_row=3, group_col="param_exit_strategy")

        # Annotation: max achievable multiplier by entry price
        ann_row = len(grouped) + 7
        ws.cell(ann_row, 1, "ℹ️  Max achievable multiplier by entry price (token capped at $1.00):")
        ws.cell(ann_row, 1).font = Font(name="Calibri", size=9, color="6B7280", italic=True)
        entries = [
            ("0.01¢ entry", "100x max"),
            ("0.02¢ entry", "50x max"),
            ("0.03¢ entry", "33x max"),
            ("0.05¢ entry", "20x max"),
            ("0.10¢ entry", "10x max"),
        ]
        for i, (ep, mx) in enumerate(entries):
            ws.cell(ann_row + 1 + i, 2, f"{ep} → {mx}")
            ws.cell(ann_row + 1 + i, 2).font = Font(name="Calibri", size=9, color="374151")

        _autowidth(ws)

    # ── Sheet: Fill Model Comparison ──────────────────────────────────────────

    def _sheet_fill_model_comparison(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("🔬 Fill Model")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "OPTIMISTIC vs REALISTIC FILL MODEL COMPARISON"
        ws["A1"].font = _title_font()

        ws["A3"] = ("The optimistic model fills if ANY historical trade occurred at your "
                    "target price. The realistic model requires sufficient volume and applies "
                    "3× queue competition + 0.5% slippage.")
        ws["A3"].font = Font(name="Calibri", size=10, color="6B7280", italic=True)
        ws.merge_cells("A3:L3")

        grouped = _group_summary(df, "fill_model")
        _write_summary_table(ws, grouped, start_row=5, group_col="fill_model")
        _autowidth(ws)

    # ── Sheet: Monthly P&L ────────────────────────────────────────────────────

    def _sheet_monthly_pnl(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("📅 Monthly P&L")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "MONTHLY P&L DISTRIBUTION"
        ws["A1"].font = _title_font()

        ws["A3"] = ("Distribution of net_pnl across all backtests, grouped by month. "
                    "Look for consistency — a strategy that only profits in 1-2 months "
                    "is less reliable than one that earns steadily.")
        ws["A3"].font = Font(name="Calibri", size=10, color="6B7280", italic=True)
        ws.merge_cells("A3:L3")

        # Pivot: avg net_pnl per (strategy, fill_model) by month
        if "total_net_pnl" not in df.columns:
            return

        top = rank_results(df, top_n=20)
        if top.empty:
            return

        # Summary stats
        row = 5
        stats_labels = [
            ("Total Configs Tested", len(df), "#,##0"),
            ("Configs Passing Filters", len(rank_results(df, top_n=len(df))), "#,##0"),
            ("Best Net P&L", df["total_net_pnl"].max(), '$#,##0.00'),
            ("Median Net P&L", df["total_net_pnl"].median(), '$#,##0.00'),
            ("Best Sharpe", df["sharpe_ratio"].max(), "0.0000"),
            ("Best Win Rate", df["win_rate"].max() / 100, "0.0%"),
        ]

        for label, val, fmt in stats_labels:
            ws.cell(row, 1, label).font = Font(name="Calibri", bold=True, size=10, color=C_SUBHEAD)
            ws.cell(row, 1).alignment = _left()
            c = ws.cell(row, 3, val)
            c.number_format = fmt
            c.font = Font(name="Calibri", bold=True, size=11, color=C_DARK_BG)
            c.alignment = _left()
            row += 1

        _autowidth(ws)

    # ── Sheet: Raw Data ───────────────────────────────────────────────────────

    def _sheet_raw(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("📋 Raw Data")
        ws.sheet_view.showGridLines = False
        ws["A1"] = f"RAW RESULTS — {len(df):,} configurations"
        ws["A1"].font = _title_font()

        headers = list(df.columns)
        row = 3
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row, ci, h.replace("param_", "").replace("_", " ").title())
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
            cell.alignment = _center()
        ws.row_dimensions[row].height = 18
        ws.freeze_panes = f"A{row+1}"

        for ri, (_, r) in enumerate(df.iterrows()):
            row += 1
            bg = PatternFill("solid", fgColor="F9FAFB") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for ci, col in enumerate(headers, start=1):
                cell = ws.cell(row, ci)
                cell.fill = bg
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = _center()
                val = r.get(col)
                if pd.notna(val):
                    cell.value = val
                else:
                    cell.value = ""

        _autowidth(ws, min_w=6, max_w=25)


# ── Utility functions ─────────────────────────────────────────────────────────

def _coerce_numerics(df: pd.DataFrame):
    num_cols = [
        "total_net_pnl", "roi_pct", "sharpe_ratio", "win_rate", "profit_factor",
        "fill_rate", "filled_trades", "total_trades", "winning_trades", "losing_trades",
        "max_drawdown", "total_capital_deployed", "total_fees", "composite_score",
    ]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)


def _compute_kpis(df: pd.DataFrame) -> list[tuple]:
    """Return list of (label, value_str, subtitle) for KPI cards."""
    _coerce_numerics(df)
    top = rank_results(df, top_n=len(df))

    kpis = [
        ("Total Configs", f"{len(df):,}", "all combinations"),
        ("Pass Filters", f"{len(top):,}", "≥10 fills, ≥5% fill rate"),
        ("Best Net P&L",
         _fmt_pnl(df["total_net_pnl"].max()) if len(df) else "$0",
         "single config"),
        ("Best Sharpe",
         f"{df['sharpe_ratio'].max():.3f}" if len(df) else "0",
         "weekly returns"),
        ("Best Win Rate",
         f"{df['win_rate'].max():.1f}%" if len(df) else "0%",
         "of filled trades"),
        ("Best ROI",
         f"{df['roi_pct'].max():.1f}%" if len(df) else "0%",
         "on deployed capital"),
    ]
    return kpis


def _group_summary(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    _coerce_numerics(df)
    agg = df.groupby(group_col).agg(
        configs=("total_net_pnl", "count"),
        avg_net_pnl=("total_net_pnl", "mean"),
        max_net_pnl=("total_net_pnl", "max"),
        avg_sharpe=("sharpe_ratio", "mean"),
        max_sharpe=("sharpe_ratio", "max"),
        avg_win_rate=("win_rate", "mean"),
        avg_fill_rate=("fill_rate", "mean"),
        avg_roi=("roi_pct", "mean"),
        avg_drawdown=("max_drawdown", "mean"),
    ).reset_index()
    return agg


def _write_summary_table(
    ws,
    df: pd.DataFrame,
    start_row: int = 3,
    group_col: Optional[str] = None,
):
    if df.empty:
        ws.cell(start_row, 1, "No data.")
        return

    headers = list(df.columns)
    header_labels = {
        "configs": "# Configs",
        "avg_net_pnl": "Avg Net P&L",
        "max_net_pnl": "Best Net P&L",
        "avg_sharpe": "Avg Sharpe",
        "max_sharpe": "Best Sharpe",
        "avg_win_rate": "Avg Win %",
        "avg_fill_rate": "Avg Fill %",
        "avg_roi": "Avg ROI %",
        "avg_drawdown": "Avg Drawdown",
    }

    row = start_row
    for ci, h in enumerate(headers, start=1):
        label = header_labels.get(h, h.replace("_", " ").title())
        cell = ws.cell(row, ci, label)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = _border()
    ws.row_dimensions[row].height = 22

    pnl_cols = {"avg_net_pnl", "max_net_pnl", "avg_drawdown"}
    pct_cols  = {"avg_win_rate", "avg_fill_rate", "avg_roi"}
    float4    = {"avg_sharpe", "max_sharpe"}

    for ri, (_, r) in enumerate(df.iterrows()):
        row += 1
        bg = PatternFill("solid", fgColor="F9FAFB") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, col in enumerate(headers, start=1):
            cell = ws.cell(row, ci)
            cell.border = _border()
            cell.fill = bg
            cell.alignment = _center()
            cell.font = _body_font()
            val = r.get(col)

            if col in pnl_cols:
                cell.value = float(val) if pd.notna(val) else 0
                cell.number_format = '$#,##0.00;($#,##0.00);"-"'
                if col != "avg_drawdown" and pd.notna(val):
                    cell.fill = PatternFill("solid", fgColor=C_GREEN_BG if val >= 0 else C_RED_BG)
                    cell.font = Font(name="Calibri", size=10, bold=True,
                                     color=C_GREEN_FG if val >= 0 else C_RED_FG)
            elif col in pct_cols:
                cell.value = (float(val) / 100) if pd.notna(val) else 0
                cell.number_format = "0.0%"
            elif col in float4:
                cell.value = round(float(val), 4) if pd.notna(val) else 0
                cell.number_format = "0.0000"
            elif col == "configs":
                cell.value = int(val) if pd.notna(val) else 0
                cell.number_format = "#,##0"
            else:
                cell.value = str(val) if pd.notna(val) else ""
